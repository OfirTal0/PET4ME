from flask import Flask, render_template, request, redirect, session, flash, send_file, jsonify
import sqlite3
from werkzeug.utils import secure_filename

from datetime import datetime,timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from io import BytesIO
import json
import os
from flask import send_from_directory
from flask import send_file
import uuid
from zoneinfo import ZoneInfo 
import urllib.parse
import math

israel_timezone = ZoneInfo("Asia/Jerusalem")

app = Flask(__name__, static_folder='static', template_folder='templates')
app.jinja_env.filters['ceil'] = lambda x: math.ceil(x)

app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB limit
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Get the directory where app.py is located
DATABASE_PATH = os.path.join(BASE_DIR, 'petforme.db')  # Make sure 'petforme.db' matches your SQLite database file name


app.secret_key = os.getenv('SECRET_KEY','default_secret_key')


def products_in_cart ():
    products_in_cart = session.get('products_in_cart', {})  # This is a dictionary {product_id: quantity}
    product_details_in_cart = []  # Create a new list to hold product details
    total_price = 0  # Initialize total price

    for product_id, quantity in products_in_cart.items():  # Loop through the dictionary
        product = query(f"SELECT * FROM products WHERE id={product_id}")

        if product:
            product_info = product[0]
            product_info = list(product_info)  # Convert product tuple to list so we can modify it
            product_info.append(quantity)  # Add the quantity as part of product info

            # Check if there is a discount (assuming the discount is in column 12, i.e., product[12])
            discount = product_info[12] if len(product_info) > 12 else 0
            price = product_info[3]  # Original price

            if discount > 0:
                # Calculate the discounted price
                discounted_price = price - (price * discount / 100)
            else:
                # No discount, use the original price
                discounted_price = price

            # Round the prices to 1 decimal point
            discounted_price = round(discounted_price, 2)
            price = round(price, 2)

            # Add the discounted price and quantity to the product info
            product_info.append(quantity * discounted_price)  # Calculate total price for this product
            product_details_in_cart.append(product_info)  # Collect product info with quantity and discount

            # Accumulate total price including discount and round to 1 decimal point
            total_price += quantity * discounted_price
            total_price = round(total_price, 2)

    return [product_details_in_cart,total_price]


@app.route('/static/<path:filename>')
def static_file(filename):
    # Serve static files from the 'static' folder with a custom cache timeout
    return send_from_directory(
        os.path.join(app.root_path, 'static'),
        filename,
        cache_timeout=timedelta(days=365)  # Set cache timeout (1 year)
    )

def query(sql: str = "", params: tuple = (), db_name=DATABASE_PATH):
    try:
        with sqlite3.connect(db_name) as conn:
            cur = conn.cursor()
            cur.execute(sql, params)  # Pass parameters to execute
            if sql.strip().lower().startswith('select'):
                return cur.fetchall()  # Fetch all results for SELECT queries
            conn.commit()
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return None
    except Exception as e:
        print(f"Error: {e}")
        return None
    
def generate_unique_filename(filename):
    ext = filename.rsplit('.', 1)[-1]  # Get the file extension
    unique_id = f"{datetime.now(israel_timezone).strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
    return f"{unique_id}.{ext}"

@app.context_processor
def inject_cart_length():
    len_product_in_cart = len(products_in_cart()[0]) if 'products_in_cart' in session else 0
    return dict(len_product_in_cart=len_product_in_cart)

@app.route('/api/download_db', methods=['GET'])
def download_db():
    return send_file('petforme.db', as_attachment=True)


@app.route('/api/cart_length', methods=['GET'])
def get_cart_length():
    len_product_in_cart = len(session['products_in_cart']) if 'products_in_cart' in session else 0
    return {'len_product_in_cart': len_product_in_cart}

import shutil

@app.route('/download_static', methods=['GET'])
def download_static():
    # Zip the static directory
    shutil.make_archive('static_files', 'zip', 'static')
    return send_file('static_files.zip', as_attachment=True)


@app.route('/', methods=['GET','POST'])
def index():
    if 'products_in_cart' not in session:
        session['products_in_cart'] = {}
    adopt = query(f"SELECT * FROM adopt ORDER BY id DESC LIMIT 1")
    products = query(f"SELECT * FROM products")
    products_on_sale = [product for product in products if product[10] == "כן"]
    first_product_on_sale = products_on_sale[0] if products_on_sale else None
    popular_products = [product for product in products if product[8] == "כן"]
    product_of_month = query("SELECT * FROM products WHERE monthlysale = 'כן' LIMIT 1")
    articles = query(f"SELECT * FROM blog")
    message = query("SELECT message FROM messages WHERE id = 1")[0][0]

    return render_template('index.html', message=message,articles =articles, popular_products=popular_products, adopt=adopt, first_product_on_sale=first_product_on_sale,product_of_month=product_of_month[0] if product_of_month else None)


@app.route('/about', methods=['GET','POST'])
def about(): 
    product_of_month = query("SELECT * FROM products WHERE monthlysale = 'כן' LIMIT 1")
    return render_template('about.html', product_of_month=product_of_month[0] if product_of_month else None )


@app.route('/new_catalog', methods=['GET','POST'])
def new_catalog(): 
    # Get selected animals and categories from the request
    animal_types = request.args.getlist('animal[]')  # Get a list of selected animal types
    categories = request.args.getlist('category[]')  # Get a list of selected categories
    sort_option = request.args.get('sort')  # Get the selected sorting option
    product_of_month = query("SELECT * FROM products WHERE monthlysale = 'כן' LIMIT 1")

    # Check if we have selected animals to display in the title
    animal_selected = None
    if animal_types:
        animal_selected = ', '.join(animal_types)

    # Start building the SQL query
    query_string = "SELECT * FROM products WHERE 1=1"
    params = []

    # If animal types were selected, add them to the query
    if animal_types:
        placeholders = ', '.join(['?'] * len(animal_types))
        query_string += f" AND animal IN ({placeholders})"
        params.extend(animal_types)

    # If categories were selected, add them to the query
    if categories:
        placeholders = ', '.join(['?'] * len(categories))
        query_string += f" AND category IN ({placeholders})"
        params.extend(categories)   

    # Add sorting
    if sort_option == "name_asc":
        products = query(query_string, params)
        products = sorted(products, key=lambda x: x[1])  # מניחים שהשם נמצא בעמודה 1
    elif sort_option == "price_asc":
        query_string += " ORDER BY price ASC"  # מיון ב-SQL
        products = query(query_string, params)
    else:
        products = query(query_string, params)
        products = sorted(products, key=lambda x: x[1])  # מניחים שהשם נמצא בעמודה 1

    product_in_cart = products_in_cart()[0]
    total_price = products_in_cart()[1]
    return render_template('new_catalog.html', product_of_month=product_of_month[0] if product_of_month else None,products=products, animal_selected=animal_selected, product_in_cart=product_in_cart, total_price=total_price)



@app.route('/show_product/<int:product_id>', methods=['GET','POST'])
def show_product(product_id):
    product = query(sql=f"SELECT * FROM products WHERE id={product_id}")

    product_in_cart = products_in_cart()[0]
    total_price = products_in_cart()[1]
    product_of_month = query("SELECT * FROM products WHERE monthlysale = 'כן' LIMIT 1")

    if product:
        return render_template('show_product.html', 
                               product_of_month=product_of_month[0] if product_of_month else None, 
                               product=product[0], 
                               product_in_cart=product_in_cart, 
                               total_price=total_price)
    else:
        return render_template('show_product.html', 
                               product_of_month=product_of_month[0] if product_of_month else None, 
                               error="Product not found")

# @app.route('/search', methods=['GET', 'POST'])
# def search():
#     text = request.args.get('text')  # Use 'GET' since it's a form with GET method
#     if text:
#         sql = f"SELECT * FROM products WHERE category LIKE '%{text}%' OR name LIKE '%{text}%' OR animal LIKE '%{text}%' OR description LIKE '%{text}%'"
#         products = query(sql)
#         return render_template('new_catalog.html', products=products)
#     else:
#         return redirect('/new_catalog')

@app.route('/contact', methods=['GET','POST'])
def contact():
    name = request.form.get('name')
    phone = request.form.get('phone')
    note = request.form.get('note',"none")
    date = datetime.now(israel_timezone).strftime('%Y-%m-%d %H:%M:%S')
    status = "new"  # Default status

    # Insert into leads table using raw SQL
    insert_sql = f"INSERT INTO leads (name, phone, date, status, note) VALUES ('{name}', '{phone}', '{date}', '{status}', '{note}')"
    query(insert_sql)  # Call existing query function for execution
    send_contact_email(name, phone, date, note)
    return redirect('/')



@app.route('/new_cart', methods=['GET', 'POST'])
def new_cart():
    product_of_month = query("SELECT * FROM products WHERE monthlysale = 'כן' LIMIT 1")
    product_details_in_cart = products_in_cart ()[0]
    total_price = products_in_cart ()[1]
    return render_template('new_cart.html',product_of_month=product_of_month[0] if product_of_month else None, products=product_details_in_cart, total_price=total_price)


@app.route('/remove_cart', methods=['POST'])
def remove_cart():
    products_in_cart = session.get('products_in_cart', {})  # Retrieve the dictionary from the session
    product_to_remove = request.form.get('remove')  # Get the product ID to remove
    
    if product_to_remove in products_in_cart:
        del products_in_cart[product_to_remove]  # Remove the product from the dictionary

    session['products_in_cart'] = products_in_cart  # Update the session with the modified cart
    return redirect('/new_cart')

@app.route('/update_cart', methods=['POST'])
def update_cart():
    product_id = request.form.get('product_id')
    action = request.form.get('action')  # Check if the action is increase or decrease
    current_quantity = int(request.form.get('quantity'))
    
    # Retrieve current cart from the session
    products_in_cart = session.get('products_in_cart', {})
    
    if product_id in products_in_cart:
        if action == "increase":
            products_in_cart[product_id] = current_quantity + 1  # Increment quantity
        elif action == "decrease" and current_quantity > 1:
            products_in_cart[product_id] = current_quantity - 1  # Decrement quantity
        else:
            # Optional: Handle case where quantity is already at 1 and can't be decreased
            flash("כמות מינימלית היא 1", "error")
    
    session['products_in_cart'] = products_in_cart  # Save updated cart back to session
    return redirect('/new_cart')

@app.route('/add_to_cart', methods=['POST'])
def add_to_cart():
    product_id = request.form.get('product_id')
    products_in_cart = session.get('products_in_cart', {})

    # Add product to the cart
    if product_id in products_in_cart:
        products_in_cart[product_id] += 1
    else:
        products_in_cart[product_id] = 1

    session['products_in_cart'] = products_in_cart

    # Get the updated cart products for the response
    # Assuming you fetch product details from the database based on the product IDs in the cart
    cart_product_details = []
    for prod_id in products_in_cart.keys():
        query_string = "SELECT * FROM products WHERE id = ?"
        product = query(query_string, [prod_id])[0]  # Assuming query returns a list
        cart_product_details.append(product)

    # Return the updated cart product details and success message as JSON
    return jsonify({
        "success": True,
        "cart_products": cart_product_details
    })


@app.route('/submit_order', methods=['POST'])
def submit_order():
    try:
        name = request.form.get('name')
        phone = request.form.get('phone')
        products_in_cart = session.get('products_in_cart', {})  # Get the cart as a dictionary
        date = datetime.now(israel_timezone).strftime('%Y-%m-%d %H:%M:%S')
        status = "new"
        address = request.form.get('address')
        note = request.form.get('note')
        customer_email = request.form.get('email')  # תוודא שהוספת שדה של אימייל בטופס


        if not products_in_cart:
            flash("העגלה ריקה, לא ניתן לשלוח הזמנה", category="error")
            return redirect('/new_cart')  # Redirect back to the cart page

        product_ids = list(products_in_cart.keys())
        quantities = list(products_in_cart.values())

        product_names_query = f"SELECT id, name FROM products WHERE id IN ({', '.join('?' for _ in product_ids)})"
        
        product_names = query(product_names_query, tuple(product_ids))

        if product_names is None:
            print("Failed to fetch product names from the database.")
            return  # Early return if there was an issue

        product_dict = {str(product[0]): product[1] for product in product_names}

        products_in_cart_str = ', '.join(
            f"{product_dict.get(product_id, 'Unknown Product')}:{quantity}"
            for product_id, quantity in zip(product_ids, quantities)
        )

        # Get the last order ID from the products_order table
        last_order_id = query("SELECT id FROM products_order ORDER BY id DESC LIMIT 1")
        
        # If no orders exist, start with 1; otherwise, increment the last ID by 1
        order_num = "PFM" + str(last_order_id[0][0] + 1 if last_order_id else 1)

        # Insert the new order into the database
        query(f"""
            INSERT INTO products_order (order_num, date, name, phone, products_order, status, address)
            VALUES ('{order_num}', '{date}', '{name}', '{phone}', '{products_in_cart_str}', '{status}', '{address}')
        """)

        # Send the order email
        send_order_email(name, phone, products_in_cart_str, date, address, order_num,note)

        send_confirmation_email(customer_email, order_num, name, phone, address, note, products_in_cart_str)

        # Clear the cart after submission
        session.pop('products_in_cart', None)
        whatsapp_message = f"""
        שלום, אני רוצה להכניס הזמנה. להלן הפרטים:
        שם: {name}
        טלפון: {phone}
        כתובת: {address}
        הערות: {note or 'אין הערות'}
        מוצרים: {products_in_cart_str}
        מספר הזמנה: {order_num}
        """
        whatsapp_url = f"https://wa.me/972509936660?text={urllib.parse.quote(whatsapp_message)}"

    
        return redirect(whatsapp_url)

    except Exception as e:
        # In case of any error, log the error and show an error message
        flash(f"משהו השתבש, נסה שוב מאוחר יותר. {str(e)}", category="error")
        return redirect('/new_cart')
    

@app.route('/blog', methods=['GET', 'POST'])
def blog():
    product_of_month = query("SELECT * FROM products WHERE monthlysale = 'כן' LIMIT 1")
    articles = query(f"SELECT * FROM blog")
    return render_template('blog.html', product_of_month=product_of_month[0] if product_of_month else None,articles= articles)

@app.route('/article/<int:id>', methods=['GET'])
def article(id):
    product_of_month = query("SELECT * FROM products WHERE monthlysale = 'כן' LIMIT 1")
    # You should query for the article by the given id and pass the data to the template
    article = query(f"SELECT * FROM blog WHERE id = {id}")
    return render_template('article.html', product_of_month=product_of_month[0] if product_of_month else None,article= article[0])

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = "ari"
        password = "ari123"
        user_username = request.form.get('username') 
        user_password = request.form.get('password')  # Corrected key
        if username == user_username and password == user_password:
            return admin_new()
    
    return render_template('login.html')

@app.route('/admin_new', methods=['GET','POST'])
def admin_new(): 
    message = query("SELECT message FROM messages WHERE id = 1")[0][0]
    leads = query("SELECT * FROM leads")
    product_orders = query("SELECT * FROM products_order")
    # Fetch products without sorting
    products = query("SELECT * FROM products")
    # Sort products by name in Python (assuming product_name is in index 1)
    products_sorted = sorted(products, key=lambda x: x[1])
    costumers = query("SELECT * FROM costumer_club")
    articles = query("SELECT * FROM blog")    
    return render_template('admin_new.html', message=message,leads=leads, product_orders=product_orders, products=products_sorted, costumers=costumers, articles=articles)

@app.route('/update_lead_status_new', methods=['POST'])
def update_lead_status_new():
    data = request.form  # Get all form data
    queries = []
    
    # Iterate over the form data to prepare the update queries
    for key, value in data.items():
        if key.startswith("status_"):
            lead_id = key.split("_")[1]  # Extract the lead ID from the key
            status = value
            queries.append(f"UPDATE leads SET status = '{status}' WHERE id = '{lead_id}'")

    # Execute all updates in a single transaction
    for query_str in queries:
        query(query_str)

    return redirect('/admin_new')

@app.route('/update_order_status_new', methods=['POST'])
def update_order_status_new():
    data = request.form  # Get all form data
    queries = []
    for key, value in data.items():
        if key.startswith("status_"):
            order_id = key.split("_")[1]  # Extract the order ID from the key
            status = value
            queries.append((status, order_id))
    for status, order_id in queries:
        query("UPDATE products_order SET status = ? WHERE id = ?", (status, order_id))

    return redirect('/admin_new')

@app.route('/update_stock_new', methods=['POST'])
def update_stock_new():
    try:
        updates = []
        for key, value in request.form.items():
            # Key format: field_productID (e.g., "animal_90")
            if "_" in key:
                field, product_id = key.split("_", 1)  # Split into field and product ID
                updates.append((field, value, product_id))

        # Apply updates to the database using the query function
        for field, value, product_id in updates:
            if not isinstance(value, str):
                value = str(value)
                value = value.strip()
            sql = f"UPDATE products SET {field} = ? WHERE id = ?"
            params = (value, product_id)
            query(sql, params)

        return redirect('/admin_new')
    
    except Exception as e:
        # Handle errors
        return jsonify({"status": "error", "message": str(e)})

@app.route('/add_product_new', methods=['POST'])
def add_product_new():
    name = request.form['name']
    description = request.form['description']
    components = request.form['components']
    category = request.form['category']
    float_price = float(request.form.get('price'))  
    popular = request.form['popular']
    price = round(float_price, 2)
    stock = int(request.form['stock'])
    animal = request.form['animal']
    weight = request.form.get('weight')
    monthly_sale = 'לא'
    sale = 'לא'
    discount = 0

    image = request.files['image']
    image_filename = 'no-image.png'  # Default image filename

    if image and allowed_file(image.filename):
        image_filename = generate_unique_filename(image.filename)
        image.save(os.path.join(app.config['UPLOAD_FOLDER'], image_filename))

    # Insert into products table
    query(f"INSERT INTO products (name, category, price, description, image, stock, weight, popular, animal, monthlysale, sale, discount, components) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
          (name, category, price, description, image_filename, stock, weight, popular, animal, monthly_sale, sale, discount, components))
    return redirect('/admin_new')


@app.route('/remove_product', methods=['POST'])
def remove_product():
    product_id = request.form.get('product_id')
    if product_id:  # Ensure the product ID exists
        query("DELETE FROM products WHERE id = ?", (product_id,))
        return '', 200  # Respond with a 200 OK status
    return 'Invalid product ID', 400  # Respond with a 400 Bad Request if ID is missing

@app.route('/add_adopt', methods=['POST'])
def add_adopt():
    name = request.form['name']
    description = request.form['description']
    type = request.form['type']
    age = request.form['age']

    # אם הגיל ריק, שים את הערך "לא ידוע"
    if not age:
        age = "לא ידוע"

    image = request.files['image']
    image_filename = 'none'  # Default image filename

    if image and allowed_file(image.filename):
        # Generate a unique filename and save the image
        image_filename = generate_unique_filename(image.filename)
        image.save(os.path.join(app.config['UPLOAD_FOLDER'], image_filename))

    # Insert into adopt table
    query(f"INSERT INTO adopt (name, type, age, description, image) VALUES (?, ?, ?, ?, ?)",
          (name, type, age, description, image_filename))
    
    return redirect('/admin_new')


@app.route('/update_article', methods=['POST'])
def update_article():
    article_id = request.form['article_id']
    name = request.form['name']
    summary = request.form['summary']
    text = request.form['text'].replace('/', '|')  # Replace '/' with '|'
    
    query(f"""
    UPDATE blog
    SET name = ?, summary = ?, text = ?
    WHERE id = ?
    """, (name, summary, text, article_id))

    return redirect('/admin_new')

@app.route('/add_article', methods=['POST'])
def add_article():
    name = request.form['name']
    summary = request.form['summary']
    text = request.form['text'].replace('/', '|')  # Replace '/' with '|'
    image = request.files['image']
    image_filename = 'article-img.png'  # Default image if none is uploaded

    # Check if an image was uploaded
    if image and allowed_file(image.filename):
        # Generate a unique filename
        image_filename = generate_unique_filename(image.filename)
        
        # Save the image in the 'blog' subdirectory
        image.save(os.path.join(app.config['UPLOAD_FOLDER'], 'blog', image_filename))
    
    # Insert article into the database
    query("""
        INSERT INTO blog (name, summary, image, text)
        VALUES (?, ?, ?, ?)
    """, (name, summary, image_filename, text))
    
    return redirect('/admin_new')

@app.route('/remove_article', methods=['POST'])
def remove_article():
    article_id = request.form['article_id']
    query("DELETE FROM blog WHERE id = ?", (article_id,))  # Add a trailing comma to make it a tuple
    return redirect('/admin_new')

@app.route('/update_message', methods=['POST'])
def update_message():
    # Get the message from the form
    new_message = request.form.get('message')

    # Update the message in the database
    if new_message:
        query("UPDATE messages SET message = ? WHERE id = 1", (new_message,))
    else:
        # If the message is empty, set it to NULL
        query("UPDATE messages SET message = NULL WHERE id = 1")

    return redirect('/admin_new')

@app.route('/customer-club-signup', methods=['POST'])
def customer_club_signup():
    # Extract form data
    name = request.form.get('name')
    phone = request.form.get('phone')
    email = request.form.get('email')
    animal_type = request.form.get('animal_type')  # New field
    date = datetime.now(israel_timezone).strftime('%Y-%m-%d %H:%M:%S')
    confirmation = request.form.get('agree_updates')  # "on" if checked

    # Convert confirmation checkbox value to a boolean
    confirmation_value = 1 if confirmation == "on" else 0

    try:
        # Use parameterized query to prevent SQL injection
        query(
            "INSERT INTO costumer_club (name, phone, email, date, confirmation, animal) VALUES (?, ?, ?, ?, ?, ?)",
            (name, phone, email, date, confirmation_value, animal_type)
        )

        # Send a confirmation email (if implemented)
        send_costumer_club_email(name, phone, email, date, confirmation_value, animal_type)

        # Show a success alert
        return "<script>alert('תודה שנרשמת למועדון הלקוחות!'); window.location.href='/';</script>"

    except Exception as e:
        print("Error:", e)  # Log the error for debugging
        return "<script>alert('שגיאה בהרשמה, אנא נסה שנית.'); window.location.href='/';</script>"

def create_table_products(table="products"):
    sql = f"CREATE TABLE IF NOT EXISTS {table} (class_id NT AUTO_INCREMENT PRIMARY KEY, name TEXT, category TEXT)"
    query(sql)

sender_email = os.environ.get("SENDER_EMAIL")
sender_password = os.environ.get("SENDER_PASSWORD")


def send_confirmation_email(customer_email, order_num, name, phone, address, note, products):
    
    sender_email = "Pets4me2024@gmail.com"
    password = "gpsq osxk onuj ghbm"    
    receiver_email = customer_email
    product_items = [item.split(':') for item in products.split(', ')]
    
    order_details = []
    for product_name, quantity in product_items:
        order_details.append((product_name, quantity))

    # Build the HTML content for the email
    order_details_html = ''.join(
        f"<tr><td>{product_name}</td><td>{quantity}</td></tr>" for product_name, quantity in order_details
    )
    
    subject = "אישור הזמנה - מספר הזמנה: " + order_num
 
    
    html_body = f"""
    <html>
    <body>

        <h2>
        שלום {name}, <br>
        תודה על הזמנתך ב-Pet4me!
        </h2>
        <p>מספר הזמנה: {order_num}</p>
        <p>טלפון: {phone}</p>
        <p>כתובת: {address}</p>
        <p>הערות: {note if note else "אין הערות"}</p>
        <p>
        אנו ניצור איתך קשר בהקדם לעדכון על מצב ההזמנה. <br>
    
        בברכה,<br>
        צוות Pet4me
    
        </p>

        <table style="border-collapse: collapse; width: 100%; margin-top:10vh;">
            <thead>
                <tr>
                    <th style="border: 1px solid #ddd; padding: 8px;">שם המוצר</th>
                    <th style="border: 1px solid #ddd; padding: 8px;">כמות</th>
                </tr>
            </thead>
            <tbody>
                {order_details_html}
            </tbody>
        </table>
    </body>
    </html>
    """
    
    # Create message
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    msg.attach(MIMEText(html_body, 'html'))  # HTML version

    try:
        # Connect to the SMTP server
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, password)
            server.send_message(msg)
            print("customer confirmation email sent successfully.")
    except Exception as e:
        print(f"Failed to send email: {e}")

def send_order_email(name, phone, products, order_date, address,order_num,note):
    # Email configuration
    sender_email = "Pets4me2024@gmail.com"
    sender_password = "gpsq osxk onuj ghbm"
    receiver_email = "Pets4me2024@gmail.com"

    # Parse the products string into a list of tuples (product_id, quantity)
    product_items = [item.split(':') for item in products.split(', ')]
    
    order_details = []
    for product_name, quantity in product_items:
        order_details.append((product_name, quantity))

    # Build the HTML content for the email
    order_details_html = ''.join(
        f"<tr><td>{product_name}</td><td>{quantity}</td></tr>" for product_name, quantity in order_details
    )

    html_body = f"""
    <html>
    <body>
        <h2>הזמנה חדשה נכנסה</h2>
        <p>מספר הזמנה: {order_num}</p>
        <p>תאריך: {order_date}</p>
        <p>שם המזמין: {name}</p>
        <p>טלפון: {phone}</p>
        <p>כתובת: {address}</p>
        <p>הערות {note}</p>

        <table style="border-collapse: collapse; width: 100%;">
            <thead>
                <tr>
                    <th style="border: 1px solid #ddd; padding: 8px;">שם המוצר</th>
                    <th style="border: 1px solid #ddd; padding: 8px;">כמות</th>
                </tr>
            </thead>
            <tbody>
                {order_details_html}
            </tbody>
        </table>
    </body>
    </html>
    """

    # Create the email message
    msg = MIMEMultipart('alternative')  # Set the email to send both plain text and HTML
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = f"New PET4ME Order Num {order_num}"

    # Attach both plain text and HTML versions
    msg.attach(MIMEText("This email requires HTML support.", 'plain'))  # Fallback for non-HTML email clients
    msg.attach(MIMEText(html_body, 'html'))  # HTML version

    try:
        # Connect to the SMTP server
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
            print("Order confirmation email sent successfully.")
    except Exception as e:
        print(f"Failed to send email: {e}")

def send_costumer_club_email(name, phone, email, date, confirmation, animal_type):
    # Email configuration
    sender_email = "Pets4me2024@gmail.com"
    sender_password = "gpsq osxk onuj ghbm"
    receiver_email = "Pets4me2024@gmail.com"
    
   
    html_body = f"""
    <html>
    <body>
        <h2>הצטרפות למועדון לקוחות</h2>
        <p>תאריך: {date}</p>
        <p>שם מלא: {name}</p>
        <p>טלפון: {phone}</p>
        <p>אימייל: {email}</p>
        <p>סוג החיה ברשותי:  {animal_type}</p>
        <p>אישר קבלת פרטים {confirmation}</p>
    </body>
    </html>
    """

    # Create the email message
    msg = MIMEMultipart('alternative')  # Set the email to send both plain text and HTML
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = "New PET4ME Costumer Club"

    # Attach both plain text and HTML versions
    msg.attach(MIMEText("This email requires HTML support.", 'plain'))  # Fallback for non-HTML email clients
    msg.attach(MIMEText(html_body, 'html'))  # HTML version

    try:
        # Connect to the SMTP server
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
            print("costumer club requset email sent successfully.")
    except Exception as e:
        print(f"Failed to send email: {e}")



def send_contact_email(name, phone, date,note):
    # Email configuration
    sender_email = "Pets4me2024@gmail.com"
    sender_password = "gpsq osxk onuj ghbm"
    receiver_email = "Pets4me2024@gmail.com"
    
   
    html_body = f"""
    <html>
    <body>
        <h2>בקשה ליצירת קשר</h2>
        <p>תאריך: {date}</p>
        <p>שם מלא: {name}</p>
        <p>טלפון: {phone}</p>
        <p>הערות: {note}</p>

    </body>
    </html>
    """

    # Create the email message
    msg = MIMEMultipart('alternative')  # Set the email to send both plain text and HTML
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = "New PET4ME Contact Request"

    # Attach both plain text and HTML versions
    msg.attach(MIMEText("This email requires HTML support.", 'plain'))  # Fallback for non-HTML email clients
    msg.attach(MIMEText(html_body, 'html'))  # HTML version

    try:
        # Connect to the SMTP server
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
            print("contact requset email sent successfully.")
    except Exception as e:
        print(f"Failed to send email: {e}")


def create_products_table():
    sql = '''
        CREATE TABLE IF NOT EXISTS products_new (
            id INTEGER PRIMARY KEY,
            מוצר TEXT,
            ברקוד TEXT,
            כמות INTEGER,
            קבוצה TEXT,
            תיאור TEXT,
            מחיר_רכישה REAL,
            מחיר_מכירה REAL,
            Tags TEXT,
            נתונים TEXT,
            כמות_מינימלית INTEGER,
            תמונה TEXT,
            משקל_שק REAL,
            דף_בית TEXT,
            מבצע TEXT,
            אחוז_מבצע REAL
        );
    '''
    query(sql)



def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/download-data', methods=['GET'])
def download_data():

    leads = query("SELECT * FROM leads")
    product_orders = query("SELECT * FROM products_order")
    products = query("SELECT * FROM products")
    costumers = query("SELECT * FROM costumer_club")

    # Create a new Excel workbook
    wb = Workbook()

    # Add a sheet for each section
    # Leads sheet
    leads_ws = wb.active
    leads_ws.title = "Leads"
    leads_ws.append(["ID", "שם", "טלפון", "תאריך", "סטטוס"])
    for lead in leads:  # Assuming `leads` is fetched from the database
        leads_ws.append(lead)

    # Customer Club sheet
    customer_ws = wb.create_sheet(title="Customer Club")
    customer_ws.append(["ID", "שם", "טלפון", "אימייל", "תאריך", "אישור התראות", "סוג חיה"])
    for customer in costumers:  # Assuming `customers` is fetched from the database
        customer_ws.append(customer)

    # Product Orders sheet
    orders_ws = wb.create_sheet(title="Product Orders")
    orders_ws.append(["ID", "תאריך", "שם", "טלפון", "הזמנות", "סטטוס", "כתובת", "מספר הזמנה"])
    for order in product_orders:  # Assuming `product_orders` is fetched from the database
        orders_ws.append(order)

    # Stock sheet
    stock_ws = wb.create_sheet(title="Stock")
    stock_ws.append(["ID","שם המוצר","קטגוריה","מחיר","תיאור","תמונה","כמות במלאי", "משקל מוצר", "האם מוגדר פופולארי","סוג חיה","מבצע חודשי","מבצע","אחוז הנחה"])
    for product in products:  # Assuming `products` is fetched from the database
        stock_ws.append(product)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file as a download
    return send_file(output, as_attachment=True, download_name="admin_data.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# if __name__ == '__main__':
#     app.run(debug=False)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", 5000)))    